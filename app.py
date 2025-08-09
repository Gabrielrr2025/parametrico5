# Finhealth — VaR Paramétrico por Classe (sem correlação) + Export Excel com Template + Mapeamento
import streamlit as st
import pandas as pd
import numpy as np
import math, datetime
from io import BytesIO
from pandas import ExcelWriter
from openpyxl import load_workbook

st.set_page_config(page_title="Finhealth • VaR Paramétrico por Classe", page_icon="📉", layout="wide")

# ---------- CSS / UI ----------
st.markdown("""
<style>
:root { --bg:#0b1220; --panel:#0f172a; --text:#e5e7eb; --muted:#94a3b8; --border:rgba(255,255,255,.08); }
html, body, [data-testid="stAppViewContainer"]{background:linear-gradient(135deg,#0b1220,#0f172a);color:var(--text)}
.fin-card{background:#0f172a;border:1px solid var(--border);border-radius:18px;padding:16px 18px;box-shadow:0 8px 28px rgba(0,0,0,.35)}
.kpi{display:flex;flex-direction:column;gap:6px;border:1px solid var(--border);border-radius:14px;padding:14px 16px;background:rgba(255,255,255,.03)}
.kpi .l{color:var(--muted);font-size:12px}.kpi .v{font-size:28px;font-weight:700}.kpi .s{color:var(--muted);font-size:12px}
.badge{display:inline-flex;align-items:center;gap:8px;border:1px solid var(--border);padding:6px 10px;border-radius:999px;background:rgba(255,255,255,.03);font-size:12px;color:var(--muted)}
hr.soft{border:none;height:1px;background:var(--border);margin:8px 0 14px}
.footer{margin-top:20px;text-align:center;color:var(--muted);font-size:13px}
</style>
""", unsafe_allow_html=True)

st.markdown("## 📉 Finhealth — VaR Paramétrico por Classe (sem correlação)")
st.markdown("<div class='badge'>Sem tickers • Sem Monte Carlo • 100% por classe de ativo</div>", unsafe_allow_html=True)
st.markdown("<hr class='soft'/>", unsafe_allow_html=True)

# ---------- helpers ----------
def nice_pct(x, d=4): 
    return f"{x*100:.{d}f}%"

def validar_aloc(df):
    req = {"Classe","%PL","Vol_Anual","FPR"}
    if not req.issubset(df.columns): 
        return "Use colunas: Classe, %PL, Vol_Anual, FPR."
    if df["%PL"].sum() <= 0:
        return "A soma de %PL precisa ser > 0."
    if df["%PL"].sum() > 100.0001:
        return "A soma de %PL não pode passar de 100%."
    if (df["Vol_Anual"] < 0).any():
        return "Vol_Anual não pode ser negativa."
    return None

def var_parametrico_sem_corr(aloc, pl, z, horizonte_dias):
    vol_d = aloc["Vol_Anual"].values / math.sqrt(252.0)
    w = (aloc["%PL"].values / 100.0)
    var_pct_por_classe = z * vol_d * math.sqrt(horizonte_dias)  # decimal
    var_rs_por_classe = pl * w * var_pct_por_classe
    out = aloc.copy()
    out["VaR_%"] = var_pct_por_classe * 100
    out["VaR_R$"] = var_rs_por_classe
    var_total_rs = var_rs_por_classe.sum()
    var_total_pct = (var_total_rs / pl) if pl > 0 else 0.0
    return out, float(var_total_rs), float(var_total_pct)

def pior_cenario_por_fpr(scen_df, aloc):
    weights_fpr = aloc.groupby("FPR")["%PL"].sum() / 100.0
    registros = []
    for fpr in scen_df["FPR"].unique():
        sub = scen_df[scen_df["FPR"] == fpr]
        w = float(weights_fpr.get(fpr, 0.0))
        if w <= 0 or sub.empty:
            continue
        sub = sub.copy()
        sub["Impacto_%PL"] = w * sub["Choque"]
        pior = sub.sort_values("Impacto_%PL").iloc[0]
        registros.append({
            "FPR": fpr,
            "Cenário Utilizado": pior["Descrição"],
            "Choque": pior["Choque"],
            "Impacto_%PL": pior["Impacto_%PL"]
        })
    return pd.DataFrame(registros)

def build_field_values_map(aloc, pl, conf_label, horizonte_dias, var_total_pct, var21_pct, pior_df):
    # impactos específicos -1%
    def impacto_chq(fpr_nome, chq=-0.01):
        peso = aloc[aloc["FPR"]==fpr_nome]["%PL"].sum()/100.0
        return peso * chq
    pior_global_pct = float(pior_df["Impacto_%PL"].min()) if (pior_df is not None and not pior_df.empty) else 0.0
    var1d_pct = var_parametrico_sem_corr(aloc, pl, 1.65 if conf_label=="95%" else 2.33, 1)[2]
    values = {
        "CNPJ": None,  # preenchido depois com o input
        "Fundo": None,
        "DataRef": None,
        "PL": None,
        "Confianca": conf_label,
        "Horizonte": horizonte_dias,
        "Modelo": "Paramétrico sem correlação",
        "Var21_95_pct": nice_pct(var21_pct, 4),
        "Var1d_pct": nice_pct(var1d_pct, 4),
        "PiorStress_pct": nice_pct(pior_global_pct, 4),
        "PiorCenarioIBOV": (pior_df[pior_df["FPR"]=="IBOVESPA"]["Cenário Utilizado"].iloc[0] if (pior_df is not None and "IBOVESPA" in pior_df.get("FPR",[]).values) else "—"),
        "PiorCenarioJurosPre": (pior_df[pior_df["FPR"]=="Juros-Pré"]["Cenário Utilizado"].iloc[0] if (pior_df is not None and "Juros-Pré" in pior_df.get("FPR",[]).values) else "—"),
        "PiorCenarioCupomCambial": (pior_df[pior_df["FPR"]=="Cupom Cambial"]["Cenário Utilizado"].iloc[0] if (pior_df is not None and "Cupom Cambial" in pior_df.get("FPR",[]).values) else "—"),
        "PiorCenarioDolar": (pior_df[pior_df["FPR"]=="Dólar"]["Cenário Utilizado"].iloc[0] if (pior_df is not None and "Dólar" in pior_df.get("FPR",[]).values) else "—"),
        "PiorCenarioOutros": (pior_df[pior_df["FPR"]=="Outros"]["Cenário Utilizado"].iloc[0] if (pior_df is not None and "Outros" in pior_df.get("FPR",[]).values) else "—"),
        "ImpactoJuros_1pct": nice_pct(impacto_chq("Juros-Pré",-0.01), 4),
        "ImpactoDolar_1pct": nice_pct(impacto_chq("Dólar",-0.01), 4),
        "ImpactoIbov_1pct": nice_pct(impacto_chq("IBOVESPA",-0.01), 4),
    }
    return values

def gerar_excel_bytes(*, meta_df, var_df, pior_df=None, aloc_df=None, respostas_df=None, scen_df=None,
                      template_file=None, cell_map_df=None, default_cell_map=None, field_values_map=None):
    """
    Gera Excel:
    - Se `template_file` é enviado, preserva o template e APPENDA abas padrão.
    - Se houver `cell_map_df` (ou `default_cell_map`), preenche as células específicas com base em `field_values_map`.
    """
    buf = BytesIO()
    if template_file is not None:
        in_mem = BytesIO(template_file.getbuffer()) if hasattr(template_file, "getbuffer") else BytesIO(template_file.read())
        wb = load_workbook(in_mem)
        with ExcelWriter(buf, engine="openpyxl") as writer:
            writer.book = wb
            if meta_df is not None:      meta_df.to_excel(writer, sheet_name="Inputs", index=False)
            if var_df is not None:       var_df.to_excel(writer, sheet_name="VaR_por_Classe", index=False)
            if pior_df is not None:      pior_df.to_excel(writer, sheet_name="Stress_PiorCenario", index=False)
            if respostas_df is not None: respostas_df.to_excel(writer, sheet_name="Respostas_CVM_B3", index=False)
            if aloc_df is not None:      aloc_df.to_excel(writer, sheet_name="Alocacao_Entrada", index=False)
            if scen_df is not None:      scen_df.to_excel(writer, sheet_name="Cenarios", index=False)
            writer._save()
        # preencher células
        mapping = cell_map_df if (cell_map_df is not None and not cell_map_df.empty) else default_cell_map
        if mapping is not None and not mapping.empty:
            out2 = BytesIO(buf.getvalue())
            wb2 = load_workbook(out2)
            for _, r in mapping.iterrows():
                sh, cell, field = str(r["Sheet"]), str(r["Cell"]), str(r["Field"])
                if sh not in wb2.sheetnames: 
                    continue
                ws = wb2[sh]
                if field_values_map and field in field_values_map:
                    val = field_values_map[field]
                else:
                    val = None
                try:
                    ws[cell].value = val
                except Exception:
                    pass
            final_buf = BytesIO()
            wb2.save(final_buf)
            final_buf.seek(0)
            return final_buf.getvalue()
        buf.seek(0)
        return buf.getvalue()
    else:
        with ExcelWriter(buf, engine="openpyxl") as writer:
            if meta_df is not None:      meta_df.to_excel(writer, sheet_name="Inputs", index=False)
            if var_df is not None:       var_df.to_excel(writer, sheet_name="VaR_por_Classe", index=False)
            if pior_df is not None:      pior_df.to_excel(writer, sheet_name="Stress_PiorCenario", index=False)
            if respostas_df is not None: respostas_df.to_excel(writer, sheet_name="Respostas_CVM_B3", index=False)
            if aloc_df is not None:      aloc_df.to_excel(writer, sheet_name="Alocacao_Entrada", index=False)
            if scen_df is not None:      scen_df.to_excel(writer, sheet_name="Cenarios", index=False)
        buf.seek(0)
        return buf.getvalue()

# ---------- mapeamento padrão (detectado) ----------
DEFAULT_CELL_MAP = pd.DataFrame([
    # Sheet, Cell, Field
    ["Informe Perfil Mensal","A9","CNPJ"],
    ["Informe Perfil Mensal","C7","Var21_95_pct"],
    ["Informe Perfil Mensal","J7","Var1d_pct"],
    ["Informe Perfil Mensal","E7","PiorCenarioIBOV"],
    ["Informe Perfil Mensal","F7","PiorCenarioJurosPre"],
    ["Informe Perfil Mensal","G7","PiorCenarioCupomCambial"],
    ["Informe Perfil Mensal","H7","PiorCenarioDolar"],
    ["Informe Perfil Mensal","I7","PiorCenarioOutros"],
    ["Informe Perfil Mensal","L7","ImpactoJuros_1pct"],
    ["Informe Perfil Mensal","M7","ImpactoDolar_1pct"],
    ["Informe Perfil Mensal","N7","ImpactoIbov_1pct"],
], columns=["Sheet","Cell","Field"])

# ---------- layout: dois painéis ----------
left, right = st.columns([1.02, 2.0])

with left:
    st.markdown("### 🧭 Painel de Controles")

    with st.expander("📋 Dados do Fundo", expanded=True):
        cnpj = st.text_input("CNPJ do Fundo *", placeholder="00.000.000/0001-00")
        nome_fundo = st.text_input("Nome do Fundo *", placeholder="Fundo XPTO")
        data_ref = st.date_input("Data de Referência *", value=datetime.date.today())
        pl = st.number_input("Patrimônio Líquido (R$) *", min_value=0.0, value=0.0, step=1000.0, format="%.2f")

    with st.expander("⚙️ Parâmetros de VaR", expanded=True):
        horizonte_dias = st.selectbox("Horizonte (dias)", [1, 10, 21], index=2,
                                      help="A CVM/B3 costuma pedir 21d para algumas respostas.")
        conf_label = st.selectbox("Nível de confiança", ["95%", "99%"], help="Delta-Normal sem correlação.")
        z = 1.65 if conf_label == "95%" else 2.33

    with st.expander("📊 Alocação por Classe", expanded=True):
        st.caption("Vol_Anual em **decimal** (ex.: 0.25 = 25% a.a.). FPR = fator primitivo (IBOVESPA, Juros-Pré, Dólar, Cupom Cambial, Outros).")
        base = pd.DataFrame({
            "Classe": ["Ações (Ibovespa)","Juros-Pré","Câmbio (Dólar)","Cupom Cambial","Crédito Privado","Multimercado","Ativos Estrangeiros","Outros"],
            "%PL":   [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0],
            "Vol_Anual": [0.25,0.08,0.15,0.12,0.05,0.18,0.16,0.10],
            "FPR": ["IBOVESPA","Juros-Pré","Dólar","Cupom Cambial","Outros","Outros","Outros","Outros"]
        })
        aloc = st.data_editor(base, num_rows="dynamic", use_container_width=True, key="aloc_editor")

    with st.expander("🧨 Cenários de Estresse (por FPR)", expanded=True):
        st.caption("Choque em **variação do fator**. Impacto no PL = (peso agregado do FPR) × Choque.")
        scen_base = pd.DataFrame({
            "FPR":["IBOVESPA","Juros-Pré","Cupom Cambial","Dólar","Outros"],
            "Descrição":["Queda de 15% no IBOV","Alta de 200 bps (≈-2%)","Queda de 1% no cupom","Queda de 5% no USD/BRL","Queda de 3% em outros"],
            "Choque":[-0.15,-0.02,-0.01,-0.05,-0.03]
        })
        scen = st.data_editor(scen_base, num_rows="dynamic", use_container_width=True, key="scen_editor")

    with st.expander("📦 Excel (template e mapeamento de células)", expanded=True):
        st.caption("Envie opcionalmente um **template .xlsx**. Abas com resultados serão anexadas.")
        tpl = st.file_uploader("Template Excel (opcional)", type=["xlsx"], key="tpl_excel")
        st.caption("Mapeamento de células (opcional) — CSV com colunas: Sheet,Cell,Field. Se não enviar, uso o mapeamento padrão do seu template.")
        cellmap_file = st.file_uploader("Mapa de células (opcional, CSV)", type=["csv"], key="cell_map")

    campos_ok = bool(cnpj.strip() and nome_fundo.strip() and pl > 0)
    st.button("🚀 Calcular", type="primary", use_container_width=True, key="calc_btn")

with right:
    st.markdown("### 📈 Resultados")
    if not campos_ok:
        st.info("Preencha CNPJ, Nome do Fundo e PL no painel esquerdo.")
    else:
        err = validar_aloc(aloc)
        if err:
            st.error(err)
        else:
            # VaR por classe e total (Delta-Normal sem correlação)
            var_df, var_total_rs, var_total_pct = var_parametrico_sem_corr(aloc, pl, z, horizonte_dias)

            # KPIs
            k1,k2,k3,k4 = st.columns(4)
            with k1:
                st.markdown(f"<div class='kpi'><div class='l'>VaR ({conf_label} / {horizonte_dias}d)</div><div class='v'>{var_total_pct*100:.2f}%</div><div class='s'>Delta-Normal (sem correlação)</div></div>", unsafe_allow_html=True)
            with k2:
                st.markdown(f"<div class='kpi'><div class='l'>VaR (R$)</div><div class='v'>R$ {var_total_rs:,.0f}</div><div class='s'>Perda potencial</div></div>", unsafe_allow_html=True)
            with k3:
                st.markdown(f"<div class='kpi'><div class='l'>Classes</div><div class='v'>{(aloc['%PL']>0).sum()}</div><div class='s'>em uso</div></div>", unsafe_allow_html=True)
            with k4:
                st.markdown(f"<div class='kpi'><div class='l'>Soma %PL</div><div class='v'>{aloc['%PL'].sum():.1f}%</div><div class='s'>deve ≤ 100%</div></div>", unsafe_allow_html=True)

            st.markdown("#### 📉 VaR por Classe")
            st.dataframe(var_df.style.format({"%PL":"{:.1f}%", "Vol_Anual":"{:.2%}", "VaR_%":"{:.2f}", "VaR_R$":"R$ {:,.0f}"}), use_container_width=True)

            # Estresse: pior cenário por FPR
            pior = pior_cenario_por_fpr(scen, aloc)
            st.markdown("#### ⚠️ Estresse — Pior cenário por FPR")
            if pior.empty:
                st.info("Mapeie FPR nas classes e/ou edite a tabela de cenários.")
            else:
                show = pior.copy()
                show["Choque"] = show["Choque"].map(lambda x: f"{x:+.1%}")
                show["Impacto_%PL"] = show["Impacto_%PL"].map(lambda x: f"{x:+.2%}")
                st.dataframe(show, use_container_width=True)

            # Respostas CVM/B3 (21d, 95% para a primeira questão)
            z95 = 1.65
            var21_df, var21_rs, var21_pct = var_parametrico_sem_corr(aloc, pl, z95, 21)
            pior_global_pct = float(pior["Impacto_%PL"].min()) if not pior.empty else 0.0

            def impacto_chq(fpr_nome, chq=-0.01):
                peso = aloc[aloc["FPR"]==fpr_nome]["%PL"].sum()/100.0
                return peso * chq

            ans = pd.DataFrame({
                "Pergunta": [
                    "Qual é o VAR (Valor de risco) de um dia como percentual do PL calculado para 21 dias úteis e 95% de confiança?",
                    "Qual classe de modelos foi utilizada para o cálculo do VAR reportado na questão anterior?",
                    "FPR IBOVESPA — cenário utilizado",
                    "FPR Juros-Pré — cenário utilizado",
                    "FPR Cupom Cambial — cenário utilizado",
                    "FPR Dólar — cenário utilizado",
                    "FPR Outros — cenário utilizado",
                    "Qual a variação diária percentual esperada para o valor da cota?",
                    "Qual a variação diária percentual esperada para o valor da cota no pior cenário de estresse definido pelo seu administrador?",
                    "Variação diária % esperada do patrimônio com -1% na taxa anual de juros (pré)",
                    "Variação diária % esperada do patrimônio com -1% na taxa de câmbio (US$/Real)",
                    "Variação diária % esperada do patrimônio com -1% no preço das ações (IBOVESPA)"
                ],
                "Resposta": [
                    nice_pct(var21_pct, 4),
                    "Paramétrico (Delta-Normal) sem correlação",
                    (pior[pior["FPR"]=="IBOVESPA"]["Cenário Utilizado"].iloc[0] if "IBOVESPA" in pior["FPR"].values else "—"),
                    (pior[pior["FPR"]=="Juros-Pré"]["Cenário Utilizado"].iloc[0] if "Juros-Pré" in pior["FPR"].values else "—"),
                    (pior[pior["FPR"]=="Cupom Cambial"]["Cenário Utilizado"].iloc[0] if "Cupom Cambial" in pior["FPR"].values else "—"),
                    (pior[pior["FPR"]=="Dólar"]["Cenário Utilizado"].iloc[0] if "Dólar" in pior["FPR"].values else "—"),
                    (pior[pior["FPR"]=="Outros"]["Cenário Utilizado"].iloc[0] if "Outros" in pior["FPR"].values else "—"),
                    # var 1d ao nível selecionado na UI
                    nice_pct(var_parametrico_sem_corr(aloc, pl, 1.65 if conf_label=="95%" else 2.33, 1)[2], 4),
                    nice_pct(pior_global_pct, 4),
                    nice_pct(impacto_chq("Juros-Pré",-0.01), 4),
                    nice_pct(impacto_chq("Dólar",-0.01), 4),
                    nice_pct(impacto_chq("IBOVESPA",-0.01), 4),
                ]
            })
            st.markdown("#### 🏛️ Respostas CVM/B3")
            st.dataframe(ans, use_container_width=True)

            # --------- Excel download (template + cell map) ---------
            st.markdown("#### 📥 Download")
            meta_df = pd.DataFrame({
                "Campo":["CNPJ","Fundo","Data de Referência","PL (BRL)","Confiança","Horizonte (d)","Modelo"],
                "Valor":[cnpj, nome_fundo, data_ref.strftime("%d/%m/%Y"), pl, conf_label, horizonte_dias, "Paramétrico sem correlação"]
            })

            # ler mapeamento enviado (se houver), senão usar default
            cell_map_df = None
            try:
                if cellmap_file is not None and cellmap_file.name.lower().endswith(".csv"):
                    cell_map_df = pd.read_csv(cellmap_file)
            except Exception:
                cell_map_df = None

            # construir mapa de valores (para preencher células)
            field_values = build_field_values_map(aloc, pl, conf_label, horizonte_dias, var_total_pct, var21_pct, pior)
            field_values["CNPJ"] = cnpj
            field_values["Fundo"] = nome_fundo
            field_values["DataRef"] = data_ref.strftime("%d/%m/%Y")
            field_values["PL"] = pl

            excel_bytes = gerar_excel_bytes(
                meta_df=meta_df,
                var_df=var_df,
                pior_df=pior if not pior.empty else None,
                aloc_df=aloc,
                respostas_df=ans,
                scen_df=scen,
                template_file=tpl,
                cell_map_df=cell_map_df,
                default_cell_map=DEFAULT_CELL_MAP,
                field_values_map=field_values
            )

            st.download_button(
                "Relatório (XLSX)",
                data=excel_bytes,
                file_name=f"relatorio_var_cvm_{nome_fundo.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

st.markdown("<div class='footer'>Feito com ❤️ Finhealth</div>", unsafe_allow_html=True)
